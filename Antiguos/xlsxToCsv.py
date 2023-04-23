#Convert file excel.xlsx to csv

import csv
import openpyxl

wb = openpyxl.load_workbook('FilesDS/salud.xlsx')
sheet = wb.get_sheet_by_name('salud')

csvFile = open('excel.csv', 'w', newline='')
csvWriter = csv.writer(csvFile)

for row in range(1, sheet.max_row + 1):
    rowData = []
    for col in range(1, sheet.max_column + 1):
        rowData.append(sheet.cell(row=row, column=col).value)
    csvWriter.writerow(rowData)

csvFile.close()

# Path: csvToXlsx.py
#Convert file excel.csv to xlsx


